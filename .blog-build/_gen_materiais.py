#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Gera os 4 materiais ricos (lead magnets) do SaudeCRM."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = os.path.dirname(os.path.abspath(__file__))
AZUL = "0055FE"; AZUL_CLARO = "EEF3FF"; CINZA = "5A6472"; ESCURO = "0D1117"
WHITE = "FFFFFF"

def style_header_row(ws, row, ncols, fill=AZUL, color=WHITE):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=True, color=color, size=10, name="Calibri")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color="BEC8D8"))

def titulo(ws, texto, sub):
    ws["A1"] = texto; ws["A1"].font = Font(bold=True, size=16, color=AZUL, name="Calibri")
    ws["A2"] = sub; ws["A2"].font = Font(size=10, color=CINZA, italic=True)
    ws["A3"] = "SaudeCRM — CRM com WhatsApp para clínicas · lp.saudecrm.com"
    ws["A3"].font = Font(size=8, color=CINZA)

# ───────────────────────────────────────────────────────────────────
# 1) PLANILHA DE ATENDIMENTO (controle de leads + follow-up)
# ───────────────────────────────────────────────────────────────────
def planilha_atendimento():
    wb = Workbook()
    ws = wb.active; ws.title = "Controle de Leads"
    titulo(ws, "Planilha de Atendimento e Follow-up", "Controle de leads da clínica — do primeiro contato ao paciente fechado")
    headers = ["Data do contato","Nome do lead","Telefone/WhatsApp","Canal de origem","Procedimento de interesse",
               "Etapa","Próxima ação","Data do follow-up","Responsável","Status","Valor estimado (R$)","Observações"]
    r0 = 5
    for i,h in enumerate(headers,1): ws.cell(row=r0,column=i,value=h)
    style_header_row(ws, r0, len(headers))
    # exemplos
    exemplos = [
        ["01/06/2026","Maria Souza","(11) 98888-0000","Instagram","Implante","Em contato","Enviar valores","03/06/2026","Recepção","Quente","4500","Pediu parcelamento"],
        ["01/06/2026","João Lima","(11) 97777-0000","Google Ads","Limpeza","Consulta agendada","Confirmar consulta","02/06/2026","Recepção","Agendado","300","Primeira vez"],
        ["02/06/2026","Ana Paula","(11) 96666-0000","Indicação","Clareamento","Novo","Fazer 1º contato","02/06/2026","Recepção","Novo","800","Indicada pela Maria"],
    ]
    for ri,row in enumerate(exemplos, r0+1):
        for ci,val in enumerate(row,1): ws.cell(row=ri,column=ci,value=val)
    # validações (listas)
    dv_etapa = DataValidation(type="list", formula1='"Novo,Em contato,Consulta agendada,Compareceu,Tratamento fechado,Perdido"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"Novo,Quente,Morno,Frio,Agendado,Fechado,Perdido"', allow_blank=True)
    dv_canal = DataValidation(type="list", formula1='"Instagram,Google Ads,Facebook,Site,WhatsApp,Indicação,Outro"', allow_blank=True)
    ws.add_data_validation(dv_etapa); ws.add_data_validation(dv_status); ws.add_data_validation(dv_canal)
    dv_canal.add(f"D{r0+1}:D200"); dv_etapa.add(f"F{r0+1}:F200"); dv_status.add(f"J{r0+1}:J200")
    widths = [14,20,20,16,22,18,22,16,14,12,16,30]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{r0+1}"
    # Aba 2 — Resumo
    ws2 = wb.create_sheet("Resumo")
    titulo(ws2, "Resumo do mês", "Acompanhe os números do seu atendimento")
    dados = [
        ("Leads recebidos no mês", '=COUNTA(\'Controle de Leads\'!B6:B200)'),
        ("Consultas agendadas", '=COUNTIF(\'Controle de Leads\'!F6:F200,"Consulta agendada")'),
        ("Tratamentos fechados", '=COUNTIF(\'Controle de Leads\'!F6:F200,"Tratamento fechado")'),
        ("Leads perdidos", '=COUNTIF(\'Controle de Leads\'!F6:F200,"Perdido")'),
        ("Taxa de conversão (fechado / leads)", '=IFERROR(B8/B5,0)'),
        ("Valor estimado em negociação (R$)", '=SUM(\'Controle de Leads\'!K6:K200)'),
    ]
    r = 5
    for label,formula in dados:
        ws2.cell(row=r,column=1,value=label).font = Font(bold=True, size=10)
        ws2.cell(row=r,column=2,value=formula)
        r += 1
    ws2.cell(row=9,column=2).number_format = "0.0%"
    ws2.column_dimensions["A"].width = 38; ws2.column_dimensions["B"].width = 18
    ws2.cell(row=r+2,column=1,value="Cansou de planilha? O SaudeCRM faz isso automático, integrado ao WhatsApp.").font = Font(italic=True, color=AZUL, size=10)
    wb.save(os.path.join(OUT,"planilha-de-atendimento-saudecrm.xlsx"))
    print("OK planilha-de-atendimento-saudecrm.xlsx")

# ───────────────────────────────────────────────────────────────────
# 2) PLANILHA DE MÉTRICAS / INDICADORES
# ───────────────────────────────────────────────────────────────────
def planilha_metricas():
    wb = Workbook(); ws = wb.active; ws.title = "Indicadores"
    titulo(ws, "Painel de Indicadores da Clínica", "Preencha os campos em azul e acompanhe seus números mês a mês")
    headers = ["Indicador","Como calcular","Jan","Fev","Mar","Abr","Mai","Jun"]
    r0 = 5
    for i,h in enumerate(headers,1): ws.cell(row=r0,column=i,value=h)
    style_header_row(ws, r0, len(headers))
    linhas = [
        ("Investimento em tráfego (R$)","Total gasto em anúncios"),
        ("Leads recebidos","Soma de todos os contatos novos"),
        ("Custo por lead (R$)","Investimento ÷ Leads"),
        ("Consultas agendadas","Leads que marcaram consulta"),
        ("Taxa lead → consulta","Consultas ÷ Leads"),
        ("Comparecimentos","Quantos compareceram"),
        ("Taxa de no-show","1 - (Comparecimentos ÷ Consultas)"),
        ("Tratamentos fechados","Quantos fecharam tratamento"),
        ("Taxa de conversão final","Fechados ÷ Leads"),
        ("Faturamento (R$)","Receita do mês"),
        ("Ticket médio (R$)","Faturamento ÷ Fechados"),
        ("Custo por paciente (R$)","Investimento ÷ Fechados"),
    ]
    for ri,(ind,calc) in enumerate(linhas, r0+1):
        ws.cell(row=ri,column=1,value=ind).font = Font(bold=True,size=10)
        ws.cell(row=ri,column=2,value=calc).font = Font(size=9,color=CINZA,italic=True)
        for col in range(3,9):  # meses editáveis em azul claro
            cell = ws.cell(row=ri,column=col)
            cell.fill = PatternFill("solid", fgColor=AZUL_CLARO)
            cell.border = Border(bottom=Side(style="hair",color="D4DBE8"),right=Side(style="hair",color="D4DBE8"))
    # fórmulas exemplo na coluna C (Jan): custo por lead, taxas etc. (linhas relativas)
    # r0+1=invest, +2=leads, +3=cpl, +4=consultas, +5=taxa l→c, +6=compareceu, +7=noshow, +8=fechados, +9=conv, +10=fat, +11=ticket, +12=cpp
    for col in range(3,9):
        L = get_column_letter(col)
        ws.cell(row=r0+3,column=col,value=f"=IFERROR({L}{r0+1}/{L}{r0+2},0)")  # CPL
        ws.cell(row=r0+5,column=col,value=f"=IFERROR({L}{r0+4}/{L}{r0+2},0)").number_format="0.0%"  # l→c
        ws.cell(row=r0+7,column=col,value=f"=IFERROR(1-{L}{r0+6}/{L}{r0+4},0)").number_format="0.0%"  # noshow
        ws.cell(row=r0+9,column=col,value=f"=IFERROR({L}{r0+8}/{L}{r0+2},0)").number_format="0.0%"  # conv
        ws.cell(row=r0+11,column=col,value=f"=IFERROR({L}{r0+10}/{L}{r0+8},0)")  # ticket
        ws.cell(row=r0+12,column=col,value=f"=IFERROR({L}{r0+1}/{L}{r0+8},0)")  # cpp
    ws.column_dimensions["A"].width=28; ws.column_dimensions["B"].width=30
    for c in range(3,9): ws.column_dimensions[get_column_letter(c)].width=10
    ws.freeze_panes="C6"
    ws.cell(row=r0+15,column=1,value="Campos em azul = você preenche. Os demais se calculam sozinhos.").font=Font(italic=True,color=CINZA,size=9)
    ws.cell(row=r0+16,column=1,value="O SaudeCRM calcula esses indicadores automaticamente, sem digitar nada.").font=Font(italic=True,color=AZUL,size=10)
    wb.save(os.path.join(OUT,"planilha-de-metricas-saudecrm.xlsx"))
    print("OK planilha-de-metricas-saudecrm.xlsx")

planilha_atendimento()
planilha_metricas()
print("planilhas geradas")
